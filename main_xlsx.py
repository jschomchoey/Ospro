import tkinter as tk
from tkinter import *
from tkinter import ttk, filedialog
import openpyxl
import os

window = tk.Tk()
window.title("Apple Stock")
window.minsize(800, 400)
# window.geometry("865x400")
# root.resizable(False, False)


button_frame = Frame()
button_frame.pack(
    fill=tk.X,
    pady=20,
    padx=20,
)


def addProduct():
    add = Toplevel(window)
    add.resizable(False, False)
    # add.geometry("750x170")
    add.title("Add Product")

    detailFrame = Frame(add)
    detailFrame.pack(
        fill=tk.X,
        pady=10,
        padx=20,
    )
    Label(detailFrame, text="ID").pack(side="left")

    IDEntry = Entry(detailFrame)
    IDEntry.pack(side="left", fill=tk.X, expand=True, padx=10)

    Label(detailFrame, text="Name").pack(side="left")

    NameEntry = Entry(detailFrame)
    NameEntry.pack(side="left", fill=tk.X, expand=True, padx=10)

    OPTIONS = [
        "32 GB",
        "64 GB",
        "128 GB",
        "256 GB",
        "512 GB",
        "1 TB",
        "2 TB",
    ]

    Label(detailFrame, text="Storage").pack(side="left")

    variable = StringVar(add)
    variable.set(OPTIONS[0])
    storage = OptionMenu(
        detailFrame,
        variable,
        *OPTIONS,
    ).pack(side="left")

    colorFrame = Frame(add)
    colorFrame.pack(
        fill=tk.X,
        pady=5,
        padx=20,
    )

    Label(colorFrame, text="Color").pack(side="left")
    ColorEntry = Entry(colorFrame)
    ColorEntry.pack(side="left", fill=tk.X, expand=True, padx=10)

    Label(colorFrame, text="Price").pack(side="left")
    PriceEntry = Entry(colorFrame)
    PriceEntry.pack(side="left", fill=tk.X, expand=True, padx=10)

    Label(colorFrame, text="Quantity").pack(side="left")
    QuantityEntry = Entry(colorFrame)
    QuantityEntry.pack(side="left", fill=tk.X, expand=True, padx=10)

    button = tk.Button(
        add,
        text="OK",
        height=3,
        command=lambda: [
            addProductToXlsx(
                IDEntry.get(),
                NameEntry.get(),
                variable.get(),
                ColorEntry.get(),
                PriceEntry.get(),
                QuantityEntry.get(),
            ),
            add.destroy(),
        ],
    )
    button.pack(
        fill=tk.X,
        expand=True,
        pady=10,
        padx=20,
    )


def addProductToXlsx(
    productID, productName, productStorage, productColor, productPrice, productQuantity
):

    path = "product.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    last_row = sheet.max_row

    sheet.cell(row=last_row + 1, column=1).value = productID
    sheet.cell(row=last_row + 1, column=2).value = productName
    sheet.cell(row=last_row + 1, column=3).value = productStorage
    sheet.cell(row=last_row + 1, column=4).value = productColor
    sheet.cell(row=last_row + 1, column=5).value = productPrice
    sheet.cell(row=last_row + 1, column=6).value = productQuantity

    workbook.save(path)

    load_data()


def removeProduct():
    remove = Toplevel(window)
    remove.resizable(False, False)
    remove.title("Remove Product")

    detailFrame = Frame(remove)
    detailFrame.pack(
        fill=tk.X,
        pady=10,
        padx=20,
    )
    Label(detailFrame, text="ID").pack(side="left")

    IDEntry = Entry(detailFrame)
    IDEntry.pack(side="left", fill=tk.X, expand=True, padx=10)

    button = tk.Button(
        remove,
        text="OK",
        height=3,
        command=lambda: [
            removeProductFromXlsx(IDEntry.get()),
            remove.destroy(),
        ],
    )
    button.pack(
        fill=tk.X,
        expand=True,
        pady=10,
        padx=20,
    )


def removeProductFromXlsx(productID):
    path = "product.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == productID:
                sheet.delete_rows(cell.row, 1)
                break

    workbook.save(path)
    load_data()


def editProduct():
    edit = Toplevel(window)
    edit.resizable(False, False)
    edit.title("Edit Product")

    detailFrame = Frame(edit)
    detailFrame.pack(
        fill=tk.X,
        pady=10,
        padx=20,
    )
    Label(detailFrame, text="ID").pack(side="left")

    IDEntry = Entry(detailFrame)
    IDEntry.pack(side="left", fill=tk.X, expand=True, padx=10)

    Label(detailFrame, text="New Name").pack(side="left")

    NameEntry = Entry(detailFrame)
    NameEntry.pack(side="left", fill=tk.X, expand=True, padx=10)

    OPTIONS = [
        "32 GB",
        "64 GB",
        "128 GB",
        "256 GB",
        "512 GB",
        "1 TB",
        "2 TB",
    ]

    Label(detailFrame, text="New Storage").pack(side="left")

    variable = StringVar(edit)
    variable.set(OPTIONS[0])
    storage = OptionMenu(
        detailFrame,
        variable,
        *OPTIONS,
    ).pack(side="left")

    colorFrame = Frame(edit)
    colorFrame.pack(
        fill=tk.X,
        pady=5,
        padx=20,
    )

    Label(colorFrame, text="New Color").pack(side="left")
    ColorEntry = Entry(colorFrame)
    ColorEntry.pack(side="left", fill=tk.X, expand=True, padx=10)

    Label(colorFrame, text="New Price").pack(side="left")
    PriceEntry = Entry(colorFrame)
    PriceEntry.pack(side="left", fill=tk.X, expand=True, padx=10)

    Label(colorFrame, text="New Quantity").pack(side="left")
    QuantityEntry = Entry(colorFrame)
    QuantityEntry.pack(side="left", fill=tk.X, expand=True, padx=10)

    button = tk.Button(
        edit,
        text="OK",
        height=3,
        command=lambda: [
            editProductInXlsx(
                IDEntry.get(),
                NameEntry.get(),
                variable.get(),
                ColorEntry.get(),
                PriceEntry.get(),
                QuantityEntry.get(),
            ),
            edit.destroy(),
        ],
    )
    button.pack(
        fill=tk.X,
        expand=True,
        pady=10,
        padx=20,
    )


def editProductInXlsx(
    productID, productName, productStorage, productColor, productPrice, productQuantity
):

    path = "product.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == productID:
                sheet.cell(row=cell.row, column=2).value = productName
                sheet.cell(row=cell.row, column=3).value = productStorage
                sheet.cell(row=cell.row, column=4).value = productColor
                sheet.cell(row=cell.row, column=5).value = productPrice
                sheet.cell(row=cell.row, column=6).value = productQuantity
                break

    workbook.save(path)
    load_data()


def findProduct():
    find = Toplevel(window)
    find.resizable(False, False)
    find.title("Find Product")

    detailFrame = Frame(find)
    detailFrame.pack(
        fill=tk.X,
        pady=10,
        padx=20,
    )
    Label(detailFrame, text="ID").pack(side="left")

    IDEntry = Entry(detailFrame)
    IDEntry.pack(side="left", fill=tk.X, expand=True, padx=10)

    Label(detailFrame, text="Name").pack(side="left")

    NameEntry = Entry(detailFrame)
    NameEntry.pack(side="left", fill=tk.X, expand=True, padx=10)

    button = tk.Button(
        find,
        text="Find",
        height=3,
        command=lambda: [
            findProductInXlsx(IDEntry.get(), NameEntry.get()),
            find.destroy(),
        ],
    )
    button.pack(
        fill=tk.X,
        expand=True,
        pady=10,
        padx=20,
    )


def findProductInXlsx(productID, productName):
    for widget in window.winfo_children():
        if isinstance(widget, ttk.Treeview):
            widget.destroy()
    path = "product.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)

    cols = list_values[0]
    tree = ttk.Treeview(window, columns=cols, show="headings")

    for col_name in cols:
        tree.heading(col_name, text=col_name)
    tree.pack(expand=True, fill="y")

    for value_tuple in list_values[1:]:
        if (not productID or value_tuple[0] == productID) and (
            not productName or value_tuple[1] == productName
        ):
            tree.insert("", tk.END, values=value_tuple)


def sellProduct():
    sell = Toplevel(window)
    sell.resizable(False, False)
    sell.title("Sell Product")

    detailFrame = Frame(sell)
    detailFrame.pack(
        fill=tk.X,
        pady=10,
        padx=20,
    )
    Label(detailFrame, text="ID").pack(side="left")

    IDEntry = Entry(detailFrame)
    IDEntry.pack(side="left", fill=tk.X, expand=True, padx=10)

    Label(detailFrame, text="Quantity").pack(side="left")

    QuantityEntry = Entry(detailFrame)
    QuantityEntry.pack(side="left", fill=tk.X, expand=True, padx=10)

    button = tk.Button(
        sell,
        text="OK",
        height=3,
        command=lambda: [
            sellProductByID(IDEntry.get(), QuantityEntry.get()),
            sell.destroy(),
        ],
    )
    button.pack(
        fill=tk.X,
        expand=True,
        pady=10,
        padx=20,
    )


def sellProductByID(productID, quantity):
    path = "product.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == productID:
                current_quantity = sheet.cell(row=cell.row, column=6).value
                new_quantity = int(current_quantity) - int(quantity)
                if new_quantity < 0:
                    return
                sheet.cell(row=cell.row, column=6).value = new_quantity
                break

    workbook.save(path)
    load_data()


def load_data():
    for widget in window.winfo_children():
        if isinstance(widget, ttk.Treeview):
            widget.destroy()

    path = "product.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)

    cols = list_values[0]
    tree = ttk.Treeview(window, columns=cols, show="headings")

    for col_name in cols:
        tree.heading(col_name, text=col_name)
    tree.pack(expand=True, fill="y")

    for value_tuple in list_values[1:]:
        tree.insert("", tk.END, values=value_tuple)


buttons = [
    ("Sell", sellProduct),
    ("Add", addProduct),
    ("Remove", removeProduct),
    ("Find", findProduct),
    ("Edit", editProduct),
    ("Refresh", load_data),
]
for text, command in buttons:
    button = tk.Button(
        button_frame,
        text=text,
        command=command,
        height=5,
    )
    button.pack(
        side="left",
        fill=tk.X,
        expand=True,
        padx=10,
    )


load_data()
window.mainloop()
